"""
Парсер транзакционной истории Binance (XLSX формат)

Формат файла (14 колонок A-N, данные начинаются с B):
- Row 0:  пустая
- Row 1:  www.binance.com (колонка L)
- Row 2:  "История транзакций" (колонка C)
- Row 3:  пустая
- Row 4:  Имя | значение | Эл. почта | значение | Адрес | значение
- Row 5:  ID пользователя | значение | Срок(UTC+5) | значение
- Row 6-8: пустые
- Row 9:  заголовки (ID пользователя, Время, Аккаунт, Операция, Коин, Изменить, Примечание)
- Row 10+: транзакции

Колонки (0-indexed):
  B (1) → пустая (иногда номер строки)
  C (2) → User_ID
  D (3) → UTC_Time (YY-MM-DD HH:MM:SS)
  E (4) → пустая (merge)
  F (5) → Account (Spot, Funding)
  G (6) → Operation (P2P Trading, Buy, Sell, Deposit, ...)
  H (7) → пустая (merge)
  I (8) → Coin (USDT, BTC, ETH...)
  J (9) → Change ("+11.04" или "-25")
  K (10) → пустая (merge)
  L (11) → Remark (ID транзакции, примечание)
"""
import logging
import re
from datetime import datetime
from typing import List, Optional, Tuple, Dict
import openpyxl

from ..base_parser import (
    BaseParser, Transaction, AccountInfo, ExpectedTotals,
    TransactionType, CounterpartyType
)

logger = logging.getLogger(__name__)


# ── Маппинг операций Binance → TransactionType ──────────────────
OPERATION_TYPE_MAP = {
    # P2P
    "p2p trading":       TransactionType.TRANSFER_OUT,  # уточняется по знаку Change
    "p2p":               TransactionType.TRANSFER_OUT,
    # Crypto buy/sell
    "buy":               TransactionType.CRYPTO_BUY,
    "sell":              TransactionType.CRYPTO_SELL,
    "small assets exchange buy": TransactionType.CRYPTO_BUY,
    "large otc trading": TransactionType.CRYPTO_BUY,     # уточняется по знаку
    # Комиссионные возвраты
    "commission rebate": TransactionType.INCOME,
    "cashback voucher":  TransactionType.INCOME,
    # Пополнение/вывод
    "deposit":           TransactionType.INCOME,
    "withdraw":          TransactionType.EXPENSE,
    "fiat deposit":      TransactionType.INCOME,
    "fiat withdraw":     TransactionType.EXPENSE,
    # Стейкинг/доход
    "simple earn flexible interest":   TransactionType.INCOME,
    "simple earn locked rewards":      TransactionType.INCOME,
    "savings interest":  TransactionType.INCOME,
    "staking rewards":   TransactionType.INCOME,
    "staking purchase":  TransactionType.EXPENSE,
    "staking redemption": TransactionType.INCOME,
    "launchpool interest": TransactionType.INCOME,
    "eth 2.0 staking":  TransactionType.EXPENSE,
    "eth 2.0 staking rewards": TransactionType.INCOME,
    # Комиссии
    "transaction fee":   TransactionType.FEE,
    "fee":               TransactionType.FEE,
    "transaction related": TransactionType.FEE,
    # Внутренние переводы
    "transfer between main and funding wallet":  TransactionType.TRANSFER_OUT,
    "transfer between main account/futures":     TransactionType.TRANSFER_OUT,
    "transfer between spot account and funding account": TransactionType.TRANSFER_OUT,
    "main and funding account transfer":         TransactionType.TRANSFER_OUT,
    "main and sub account transfer":             TransactionType.TRANSFER_OUT,
    # Binance Pay / Crypto Box / Merchant
    "crypto box":         TransactionType.INCOME,
    "binance pay":        TransactionType.INCOME,      # уточняется по знаку
    "merchant payment":   TransactionType.EXPENSE,     # уточняется по знаку
    "c2c transfer":       TransactionType.TRANSFER_OUT,
    "send":               TransactionType.TRANSFER_OUT,
    "receive":            TransactionType.TRANSFER_IN,
    # Futures / Margin
    "realized profit and loss":  TransactionType.OTHER,   # уточняется по знаку
    "funding fee":               TransactionType.FEE,
    "insurance fund":            TransactionType.FEE,
    # Transaction (фьючерсы/маржа)
    "transaction sold":     TransactionType.CRYPTO_SELL,
    "transaction revenue":  TransactionType.INCOME,
    "transaction buy":      TransactionType.CRYPTO_BUY,
    "transaction spend":    TransactionType.EXPENSE,
    # Simple Earn
    "simple earn flexible subscription": TransactionType.EXPENSE,   # вложение
    "simple earn flexible redemption":   TransactionType.INCOME,    # вывод
    "simple earn locked subscription":   TransactionType.EXPENSE,
    "simple earn locked redemption":     TransactionType.INCOME,
    # Binance Convert
    "binance convert":    TransactionType.CRYPTO_BUY,   # уточняется по знаку
    # Merchant
    "merchant acquiring": TransactionType.INCOME,       # уточняется по знаку
    # Прочее
    "referral kickback":  TransactionType.INCOME,
    "distribution":       TransactionType.INCOME,
    "airdrop":            TransactionType.INCOME,
    "airdrop assets":     TransactionType.INCOME,
    "asset recovery":     TransactionType.OTHER,
    "token swap - redenomination/rebranding": TransactionType.OTHER,
    "token swap":         TransactionType.OTHER,
    "crypto box refund":  TransactionType.INCOME,
    # Convert
    "convert":            TransactionType.CRYPTO_BUY,   # уточняется по знаку
    "auto-invest":        TransactionType.CRYPTO_BUY,
    # Send / Receive
    "send":               TransactionType.TRANSFER_OUT,
    "receive":            TransactionType.TRANSFER_IN,
}


class BinanceParser(BaseParser):
    """
    Парсер XLSX-выгрузок истории транзакций Binance.
    Файл: "Binance-История-транзакций.xlsx" (Sheet0)

    ВАЖНО: openpyxl.load_workbook(read_only=False) — read_only=True
    не работает корректно с некоторыми Binance XLSX файлами.
    """

    def __init__(self, file_path: str):
        super().__init__(file_path)
        self.account.bank_name = "Binance"
        self.account.currency = "USDT"

    def parse(self) -> bool:
        """Основной метод парсинга XLSX"""
        try:
            logger.info(f"Парсинг Binance XLSX: {self.pdf_path}")

            # КРИТИЧНО: read_only=False, т.к. read_only=True иногда
            # не читает все строки в Binance XLSX файлах
            wb = openpyxl.load_workbook(self.pdf_path, data_only=True)

            # Используем первый лист (Sheet0 или по индексу)
            sheet = wb.worksheets[0]

            logger.info(
                f"Лист: {sheet.title}, строк: {sheet.max_row}, "
                f"колонок: {sheet.max_column}, dimensions: {sheet.dimensions}"
            )

            # Считываем все строки
            rows = []
            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, values_only=True
            ):
                rows.append(row)

            wb.close()

            if not rows or len(rows) < 2:
                self.errors.append("Файл пустой или содержит менее 2 строк")
                return False

            logger.info(f"Прочитано строк: {len(rows)}")

            # ── Шаг 1: Парсинг метаданных (строки 0-8) ────────────
            self._parse_metadata(rows[:10])

            # ── Шаг 2: Найти строку с заголовками ─────────────────
            header_row_idx = self._find_header_row(rows)
            if header_row_idx < 0:
                self.errors.append("Не найдена строка заголовков")
                return False

            logger.info(f"Заголовки найдены в строке {header_row_idx}")

            # ── Шаг 3: Определить индексы колонок из заголовков ────
            col_map = self._build_column_map(rows[header_row_idx])
            logger.info(f"Маппинг колонок: {col_map}")

            # ── Шаг 4: Парсинг транзакций ─────────────────────────
            parse_errors = 0
            for row_idx in range(header_row_idx + 1, len(rows)):
                row = rows[row_idx]
                if not row or all(cell is None for cell in row):
                    continue
                tx = self._parse_transaction_row(row, row_idx, col_map)
                if tx:
                    self.transactions.append(tx)
                else:
                    parse_errors += 1

            # Определяем period_start и period_end из транзакций
            if self.transactions and not self.account.period_start:
                dates = [t.date for t in self.transactions if t.date]
                if dates:
                    self.account.period_start = min(dates)
                    self.account.period_end = max(dates)

            logger.info(
                f"Успешно спарсено {len(self.transactions)} транзакций Binance "
                f"({parse_errors} ошибок парсинга)"
            )
            return len(self.transactions) > 0

        except Exception as e:
            logger.error(f"Ошибка парсинга Binance XLSX: {e}", exc_info=True)
            self.errors.append(f"Ошибка парсинга: {str(e)}")
            return False

    def _parse_metadata(self, rows: list) -> None:
        """
        Извлечь информацию о владельце из первых строк файла.

        Реальный формат:
          Row 4: (None, None, 'Имя', 'Ниязбек...', None, 'Эл. почта', 'email@...', None, 'Адрес', '...')
          Row 5: (None, None, 'ID пользователя', '743233966', None, 'Срок(UTC+5)', '2020-02-13 to 2026-02-15')
        """
        for row in rows:
            if not row:
                continue

            # Ищем пары ключ-значение в строке
            cells = list(row)
            for i in range(len(cells) - 1):
                cell = cells[i]
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                cell_lower = cell_str.lower()

                # Значение — следующая НЕ-None ячейка
                next_val = None
                for j in range(i + 1, min(i + 3, len(cells))):
                    if cells[j] is not None:
                        next_val = str(cells[j]).strip()
                        break

                if not next_val:
                    # Попытка извлечь значение из "Key: Value"
                    if ":" in cell_str:
                        parts = cell_str.split(":", 1)
                        cell_lower = parts[0].strip().lower()
                        next_val = parts[1].strip()

                if not next_val:
                    continue

                # Маппинг
                if cell_lower in ("имя", "name", "full name"):
                    if not self.account.owner:
                        self.account.owner = next_val
                elif cell_lower in ("id пользователя", "user id", "user_id"):
                    # Пропускаем строку заголовков (user_id → "Время" = строка заголовков)
                    if next_val and next_val.lower() not in ("время", "utc_time", "time"):
                        self.account.extra["user_id"] = next_val
                        if not self.account.account_number:
                            self.account.account_number = next_val
                elif cell_lower in ("эл. почта", "email", "e-mail"):
                    self.account.extra["email"] = next_val
                elif cell_lower in ("адрес", "address"):
                    self.account.extra["address"] = next_val
                elif "срок" in cell_lower or "period" in cell_lower:
                    # Формат: "2020-02-13 to 2026-02-15"
                    self._parse_period(next_val)

    def _parse_period(self, period_str: str) -> None:
        """Парсинг периода: '2020-02-13 to 2026-02-15'"""
        if not period_str:
            return
        # Разделители: "to", "-", "~"
        parts = re.split(r'\s+to\s+|\s*~\s*|\s*-\s*(?=\d{4})', period_str)
        if len(parts) >= 2:
            self.account.period_start = self._parse_datetime(parts[0].strip())
            self.account.period_end = self._parse_datetime(parts[-1].strip())
        elif len(parts) == 1:
            self.account.period_start = self._parse_datetime(parts[0].strip())

    def _find_header_row(self, rows: list) -> int:
        """
        Найти индекс строки с заголовками.
        Ищем строку где есть 'Время' / 'Time' / 'UTC_Time' И 'Изменить' / 'Change'
        """
        for i, row in enumerate(rows[:15]):  # ищем только в первых 15 строках
            if not row:
                continue
            row_text = " ".join(str(c or "").lower() for c in row)
            # Проверяем русские и английские варианты
            has_time = any(kw in row_text for kw in ("время", "utc_time", "time"))
            has_change = any(kw in row_text for kw in ("изменить", "change", "amount"))
            has_operation = any(kw in row_text for kw in ("операция", "operation", "type"))
            if has_time and (has_change or has_operation):
                return i

        # Резервный вариант: строка 9 (типичная для Binance)
        logger.warning("Заголовки не найдены по ключевым словам, используем строку 9")
        return 9

    def _build_column_map(self, header_row: tuple) -> Dict[str, int]:
        """
        Построить маппинг имён колонок на индексы из строки заголовков.
        Возвращает словарь: {'user_id': 2, 'time': 3, 'account': 5, ...}
        """
        col_map = {}
        if not header_row:
            return self._default_column_map()

        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            cell_str = str(cell).strip().lower()

            if cell_str in ("id пользователя", "user id", "user_id", "userid"):
                col_map["user_id"] = idx
            elif cell_str in ("время", "utc_time", "time", "date"):
                col_map["time"] = idx
            elif cell_str in ("аккаунт", "account", "account category"):
                col_map["account"] = idx
            elif cell_str in ("операция", "operation", "type"):
                col_map["operation"] = idx
            elif cell_str in ("коин", "coin", "asset", "token"):
                col_map["coin"] = idx
            elif cell_str in ("изменить", "change", "amount", "quantity"):
                col_map["change"] = idx
            elif cell_str in ("примечание", "remark", "note", "memo"):
                col_map["remark"] = idx

        # Если не все поля найдены — заполняем дефолтами
        if "time" not in col_map or "change" not in col_map:
            logger.warning(f"Неполный маппинг из заголовков: {col_map}, дополняем дефолтами")
            default = self._default_column_map()
            for key, val in default.items():
                if key not in col_map:
                    col_map[key] = val

        return col_map

    def _default_column_map(self) -> Dict[str, int]:
        """Дефолтный маппинг колонок для типичного Binance XLSX"""
        return {
            "user_id":   2,   # C
            "time":      3,   # D
            "account":   5,   # F
            "operation": 6,   # G
            "coin":      8,   # I
            "change":    9,   # J
            "remark":    11,  # L
        }

    def _parse_transaction_row(
        self,
        row: tuple,
        row_idx: int,
        col_map: Dict[str, int]
    ) -> Optional[Transaction]:
        """Парсинг строки транзакции Binance по динамическому маппингу колонок."""
        if not row or len(row) < 8:
            return None

        try:
            user_id   = self._safe_str(row, col_map.get("user_id", 2))
            date_str  = self._safe_str(row, col_map.get("time", 3))
            account   = self._safe_str(row, col_map.get("account", 5))
            operation = self._safe_str(row, col_map.get("operation", 6))
            coin      = self._safe_str(row, col_map.get("coin", 8))
            change_str= self._safe_str(row, col_map.get("change", 9))
            remark    = self._safe_str(row, col_map.get("remark", 11))

            # Пропускаем строки без даты или суммы
            if not date_str or not change_str:
                return None

            # Парсинг даты
            date = self._parse_datetime(date_str)
            if not date:
                return None

            # Парсинг суммы
            amount = self._parse_amount(change_str)
            if amount is None:
                return None

            # Маппинг типа операции
            op_lower = (operation or "").lower().strip()
            tx_type = self._resolve_tx_type(op_lower, amount)

            # Определение категории
            category = self._get_category(op_lower, coin or "")

            # Тип контрагента
            cp_type = self._resolve_counterparty_type(op_lower)

            description = f"{operation or 'Unknown'} {coin or ''}"
            if remark:
                description += f" {remark}"
            description = description.strip()

            return Transaction(
                date=date,
                amount=amount,
                type=tx_type,
                description=description,
                category=category,
                currency=coin or "USDT",
                counterparty=remark or operation or "Binance",
                counterparty_type=cp_type,
                reference=remark,
                raw_data={
                    "user_id": user_id,
                    "account": account,
                    "operation": operation,
                    "coin": coin,
                    "change": change_str,
                    "remark": remark,
                    "row_idx": row_idx,
                },
                source_row=row_idx,
            )

        except Exception as e:
            logger.debug(f"Ошибка парсинга строки {row_idx}: {e}")
            return None

    def _resolve_tx_type(self, op_lower: str, amount: float) -> TransactionType:
        """Определить тип транзакции по операции и знаку суммы."""
        # Точное совпадение
        tx_type = OPERATION_TYPE_MAP.get(op_lower)

        if tx_type is None:
            # Ищем частичное совпадение
            for key, val in OPERATION_TYPE_MAP.items():
                if key in op_lower:
                    tx_type = val
                    break

        if tx_type is None:
            tx_type = TransactionType.OTHER

        # Для P2P уточняем по знаку
        if "p2p" in op_lower:
            tx_type = TransactionType.TRANSFER_IN if amount > 0 else TransactionType.TRANSFER_OUT

        # Для типов зависящих от знака
        sign_dependent_ops = (
            "convert", "binance convert", "large otc trading",
            "binance pay", "merchant payment", "merchant acquiring",
            "realized profit and loss", "crypto box",
        )
        if op_lower in sign_dependent_ops:
            if amount > 0:
                tx_type = TransactionType.INCOME
            else:
                tx_type = TransactionType.EXPENSE

        # Для внутренних переводов
        if "transfer" in op_lower and "wallet" in op_lower:
            tx_type = TransactionType.TRANSFER_IN if amount > 0 else TransactionType.TRANSFER_OUT

        return tx_type

    def _resolve_counterparty_type(self, op_lower: str) -> CounterpartyType:
        """Определить тип контрагента по операции."""
        if "p2p" in op_lower or "c2c" in op_lower:
            return CounterpartyType.PERSON
        elif "deposit" in op_lower or "withdraw" in op_lower:
            return CounterpartyType.BANK
        elif any(kw in op_lower for kw in ("buy", "sell", "convert", "merchant")):
            return CounterpartyType.MERCHANT
        elif "transfer" in op_lower:
            return CounterpartyType.BANK  # внутренний перевод = банковская операция
        return CounterpartyType.UNKNOWN

    def _get_category(self, operation: str, coin: str) -> str:
        """Определить категорию транзакции"""
        if "p2p" in operation:
            return "P2P Trading"
        elif any(kw in operation for kw in ("transaction buy", "transaction spend")):
            return "Futures/Margin"
        elif any(kw in operation for kw in ("transaction sold", "transaction revenue")):
            return "Futures/Margin"
        elif "realized profit" in operation:
            return "Futures/Margin"
        elif any(kw in operation for kw in ("buy", "convert", "auto-invest", "binance convert")):
            return "Crypto Buy"
        elif "sell" in operation:
            return "Crypto Sell"
        elif "deposit" in operation or "fiat deposit" in operation:
            return "Deposit"
        elif "withdraw" in operation:
            return "Withdrawal"
        elif any(kw in operation for kw in ("interest", "staking", "earn", "reward", "subscription", "redemption")):
            return "Staking/Earn"
        elif any(kw in operation for kw in ("fee", "commission", "funding fee", "insurance")):
            return "Fee"
        elif "transfer" in operation:
            return "Internal Transfer"
        elif "crypto box" in operation:
            return "Crypto Box"
        elif any(kw in operation for kw in ("merchant", "binance pay", "acquiring")):
            return "Binance Pay"
        elif "asset recovery" in operation:
            return "Asset Recovery"
        elif any(kw in operation for kw in ("airdrop", "distribution")):
            return "Airdrop"
        elif "send" in operation:
            return "Send"
        elif "receive" in operation:
            return "Receive"
        return "Other"

    def _parse_datetime(self, date_str: str) -> Optional[datetime]:
        """
        Парсинг даты/времени:
        - "24-01-15 12:34:56" (YY-MM-DD)
        - "2024-01-15 12:34:56" (YYYY-MM-DD)
        - "2024-01-15T12:34:56"
        - "23-06-10 19:38:38" (реальный формат из файла)
        - "2020-02-13" (дата без времени)
        """
        if not date_str:
            return None

        date_str = str(date_str).strip()

        formats = [
            "%y-%m-%d %H:%M:%S",    # 23-06-10 19:38:38 (основной формат!)
            "%Y-%m-%d %H:%M:%S",    # 2024-01-15 12:34:56
            "%Y-%m-%dT%H:%M:%S",    # 2024-01-15T12:34:56
            "%Y-%m-%d",             # 2024-01-15
            "%y-%m-%d",             # 24-01-15
            "%d.%m.%Y %H:%M:%S",   # 15.01.2024 12:34:56
            "%d.%m.%Y",             # 15.01.2024
            "%d.%m.%y",             # 15.01.24
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue

        # Попытка с dateutil если стандартные не работают
        try:
            from dateutil import parser as dateutil_parser
            return dateutil_parser.parse(date_str)
        except Exception:
            pass

        logger.debug(f"Не удалось распарсить дату: {date_str!r}")
        return None

    def _parse_amount(self, value) -> Optional[float]:
        """
        Парсинг суммы: "+11.04", "-25", "11.04", "-0.00001234", числа
        Возвращает None если не число.
        """
        if value is None:
            return None

        # Если уже число — вернуть как есть
        if isinstance(value, (int, float)):
            return float(value)

        value_str = str(value).strip()
        if not value_str or value_str in ("-", "+", "N/A", "n/a", ""):
            return None

        # Удалить все кроме цифр, точки, запятой, знака
        cleaned = re.sub(r'[^\d.,+\-eE]', '', value_str)
        cleaned = cleaned.replace(',', '.')

        # Убрать лишние точки (оставить только последнюю)
        parts = cleaned.split('.')
        if len(parts) > 2:
            cleaned = ''.join(parts[:-1]) + '.' + parts[-1]

        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None

    def _safe_str(self, row: tuple, idx: int) -> Optional[str]:
        """Безопасно получить строковое значение из строки по индексу"""
        if idx >= len(row) or row[idx] is None:
            return None
        val = str(row[idx]).strip()
        return val if val else None

    # ── Обязательные абстрактные методы BaseParser ──────────────

    def _parse_account_info(self) -> None:
        """Реализовано в parse() через _parse_metadata"""
        pass

    def _parse_transactions(self) -> None:
        """Реализовано в parse()"""
        pass
